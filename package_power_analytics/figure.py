import openpyxl as xl
from openpyxl.chart import Reference, BarChart
import matplotlib.pyplot as plt
from docxtpl import DocxTemplate, InlineImage
# import seaborn as sns


# __________________________ГРАФИК В ЭКСЕЛЕ______________________#
# workbook = xl.load_workbook('Result_Excel/Результаты расчета.xlsx')
# sheet_1 = workbook['Sheet1']
# values = Reference(sheet_1, min_row=1, max_row=sheet_1.max_row, min_col=12, max_col=14)
# index = Reference(sheet_1, min_row=2, max_row=sheet_1.max_row, min_col=1)
# chart = BarChart()
# chart.y_axis.title = 'Суммарная оплата за электроэнергию, руб.'
# chart.x_axis.title = 'Расчетный период'
# chart.add_data(values, titles_from_data=True)
# chart.set_categories(index)
# sheet_1.add_chart(chart, 'E15')
# workbook.save('Result_Excel/Результаты расчета.xlsx')
# input_file = 'Result_Excel/Результаты расчета.xlsx'
# output_image = 'Result_Excel/'
# __________________________конец построения графика в экселе______________________#

# ______________________________ГРАФИКИ ПО ТЕКСТУ ОТЧЕТА____________________________#
# __________________________ГРАФИКИ ИСХОДНЫЕ В matplotlib______________________#

def my_subplots(template, data, name_fig):
    if len(data) > 1008:  # (3 недели)
        """На входе подается шаблон отчета template и df с индексом (ось Х) и данными (ось Y)"""
        plt.rcParams["font.family"] = "Times New Roman"
        fig, ax = plt.subplots(figsize=(7, 4.5))
        ax.plot(data, marker='.', markersize=1, color='0.5', linestyle='None', label='30-и минутные наблюденения')
        # Усредненные данные с помощью функции rolling
        data_mean_week = data.rolling(window=336, center=True).mean()
        data_48 = data_mean_week
        ax.plot(data_48, color='r', label='Усредненные 7-и дневные наблюдения')
        ax.set_ylabel(data.columns[0])
        ax.set_xlabel(data.index.name)
        ax.legend(bbox_to_anchor=(1, 1), loc='lower right')
        fig_file_name = 'Figure/' + name_fig + '.png'
        fig.savefig(fig_file_name, dpi=700, )
        image = InlineImage(template, fig_file_name)
        plt.close()
        return image
    else:
        """На входе подается шаблон отчета template и df с индексом (ось Х) и данными (ось Y)"""
        plt.rcParams["font.family"] = "Times New Roman"
        fig, ax = plt.subplots(figsize=(7, 4.5))
        fig.autofmt_xdate()
        ax.plot(data, color='black', label='30-и минутные наблюденения')
        ax.set_ylabel(data.columns[0])
        ax.set_xlabel(data.index.name)
        ax.legend(bbox_to_anchor=(1, 1), loc='lower right')
        fig_file_name = 'Figure/' + name_fig + '.png'
        fig.savefig(fig_file_name, dpi=700, )
        image = InlineImage(template, fig_file_name)
        plt.close()
        return image


def mu_plot(template, data, list_legends, name_xlabel, name_ylabel, name_fig):
    plt.rcParams["font.family"] = "Times New Roman"
    fig, ax = plt.subplots(figsize=(7, 4), )
    ax.plot(data, marker='o', mec='k', mfc='w', markersize=5, linewidth=2)
    ax.plot(data, markersize=1, color='0.5')
    plt.legend(list_legends, frameon=False)
    ax.set_xlabel(name_xlabel)
    ax.set_ylabel(name_ylabel)
    fig_file_name = 'Figure/' + name_fig + '.png'
    fig.savefig(fig_file_name, dpi=700, )
    image = InlineImage(template, fig_file_name)
    plt.close()
    return image


def my_bar(template, data, name_fig, colormap=None, ylim_min=None, ylim_max=None,
           loc='lower right', bbox_to_anchor=(1, 1), fontsize=6 ):
    """На входе подается шаблон отчета template и df с индексом (ось Х) и данными (ось Y)"""

    plt.rcParams["font.family"] = "Times New Roman"
    fig = data.plot.bar(rot=0, figsize=(7, 3.5), fontsize=8, colormap=colormap)
    fig.legend(bbox_to_anchor=bbox_to_anchor, loc=loc, fontsize=fontsize)
    fig.fontsize = 6
    plt.ylim((ylim_min, ylim_max))
    fig_file_name = 'Figure/' + name_fig + '.png'
    plt.savefig(fig_file_name, dpi=700)
    image = InlineImage(template, fig_file_name)
    plt.close()
    return image




def my_plot_limit(template, data, name_fig):
    """На входе подается шаблон отчета template и df """
    fig = data.plot(figsize=(7, 4.5))
    fig.set_xlabel('Период наблюдений')
    fig.legend(bbox_to_anchor=(1, 1), loc='lower right', fontsize=6)
    fig_file_name = 'Figure/' + name_fig + '.png'
    plt.savefig(fig_file_name, dpi=700)
    image = InlineImage(template, fig_file_name)
    plt.close()
    return image


# Круговая диаграмма
def my_pie(template, list_data, labels, name_fig):
    """На входе подается шаблон отчета template, список данных и список меток """
    fig = plt.figure(figsize=(5, 3.5))
    plt.pie(x=list_data, labels=labels, autopct='%.1f%%', labeldistance=None)
    plt.legend(bbox_to_anchor=(0, 0), loc='lower center', fontsize=6)
    fig_file_name = 'Figure/' + name_fig + '.png'
    plt.savefig(fig_file_name, dpi=700)
    image = InlineImage(template, fig_file_name)
    plt.close()
    return image
